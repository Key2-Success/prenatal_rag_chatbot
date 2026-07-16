"""
build_metrics_xlsx.py — regenerate the RAGAS-metrics-over-time spreadsheet.

Scans every eval report in eval/results/, extracts the three RAGAS answer-quality
means (faithfulness, context_precision, answer_relevancy) plus the run note, and
writes them — one row per scored run, chronological — to an .xlsx for plotting.

Design: it ALWAYS rebuilds from the reports on disk (the source of truth) rather
than appending a row. That makes it idempotent — safe to run repeatedly, and it
self-heals if a report is deleted, re-run, or added out of order. No dedup state.

Runs automatically at the end of every `eval.ragas_eval` (hooked in
write_markdown_report), and can be run by hand:

    python -m eval.build_metrics_xlsx                 # default output at repo root
    python -m eval.build_metrics_xlsx --output foo.xlsx
"""

import argparse
import glob
import re
from datetime import datetime
from pathlib import Path

from backend.app.config import PROJECT_ROOT

RESULTS_DIR = PROJECT_ROOT / "eval" / "results"
DEFAULT_OUTPUT = PROJECT_ROOT / "ragas_metrics_over_time.xlsx"

_TS_RE = re.compile(r"(\d{8})_(\d{6})")
_FLOAT_RE = re.compile(r"\d+\.\d+")
_NOTE_RE = re.compile(r"##\s*Note\s*\n+>\s*(.+)")

# Metric label variants seen across the report history (the format and the
# context_precision key name both changed over time).
_METRIC_ALIASES = {
    "faithfulness": ("faithfulness",),
    "context_precision": ("context_precision", "context precision"),
    "answer_relevancy": ("answer_relevancy", "answer relevancy", "answer relevance"),
}


def _parse_timestamp(filename: str) -> datetime | None:
    m = _TS_RE.search(filename)
    return datetime.strptime(m.group(1) + m.group(2), "%Y%m%d%H%M%S") if m else None


def _first_metric_value(text: str, names: tuple[str, ...]) -> float | None:
    """First markdown table row mentioning a metric name that holds a [0,1] float."""
    for line in text.splitlines():
        if "|" not in line:
            continue
        low = line.lower()
        if any(n in low for n in names):
            for tok in _FLOAT_RE.findall(line):
                v = float(tok)
                if 0.0 <= v <= 1.0:
                    return v
    return None


def _parse_report(path: Path) -> dict | None:
    """Extract one run's metrics, or None if the report has no RAGAS scores."""
    text = path.read_text(encoding="utf-8", errors="replace")
    if "faithfulness" not in text.lower():
        return None
    metrics = {
        key: _first_metric_value(text, aliases)
        for key, aliases in _METRIC_ALIASES.items()
    }
    # Routing-only runs mention "faithfulness" in prose but score nothing — skip.
    if all(v is None for v in metrics.values()):
        return None
    note_match = _NOTE_RE.search(text)
    return {
        "file": path.name,
        "ts": _parse_timestamp(path.name),
        "note": note_match.group(1).strip() if note_match else "",
        **metrics,
    }


def collect_rows(results_dir: Path = RESULTS_DIR) -> list[dict]:
    rows = []
    for f in sorted(results_dir.glob("*.md")):
        row = _parse_report(f)
        if row is not None:
            rows.append(row)
    rows.sort(key=lambda r: r["ts"] or datetime.min)
    for i, r in enumerate(rows, 1):
        r["run_number"] = i
    return rows


def write_xlsx(rows: list[dict], output_path: Path = DEFAULT_OUTPUT) -> Path | None:
    """Write rows to .xlsx. Returns None (with a hint) if openpyxl isn't installed."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        print(
            "  build_metrics_xlsx: openpyxl not installed — skipping. "
            "Install with `pip install openpyxl` to enable the metrics sheet."
        )
        return None

    headers = [
        "run_number", "timestamp", "faithfulness",
        "context_precision", "answer_relevancy", "note",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "RAGAS over time"
    ws.append(headers)

    fill = PatternFill("solid", fgColor="1F4E78")
    for c in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=c)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center")

    for r in rows:
        ws.append([
            r["run_number"], r["ts"], r["faithfulness"],
            r["context_precision"], r["answer_relevancy"], r["note"],
        ])
    for row in range(2, ws.max_row + 1):
        ws.cell(row=row, column=2).number_format = "yyyy-mm-dd hh:mm"
        for col in (3, 4, 5):
            ws.cell(row=row, column=col).number_format = "0.000"
    for col, width in zip("ABCDEF", (11, 18, 13, 17, 16, 70)):
        ws.column_dimensions[col].width = width
    ws.freeze_panes = "A2"

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


def refresh_metrics_xlsx(
    results_dir: Path = RESULTS_DIR, output_path: Path = DEFAULT_OUTPUT
) -> Path | None:
    """Rebuild the metrics spreadsheet from all reports. Returns the path or None."""
    return write_xlsx(collect_rows(results_dir), output_path)


def main() -> int:
    ap = argparse.ArgumentParser(description="Rebuild the RAGAS metrics spreadsheet.")
    ap.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="output .xlsx path")
    ap.add_argument("--results-dir", type=Path, default=RESULTS_DIR, help="reports dir")
    args = ap.parse_args()
    rows = collect_rows(args.results_dir)
    out = write_xlsx(rows, args.output)
    if out is not None:
        print(f"Wrote {out} with {len(rows)} scored runs.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
